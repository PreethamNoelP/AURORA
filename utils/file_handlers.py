"""
File handling utilities for AURORA platform
"""

import os
import tempfile
import pandas as pd
from pdfminer.high_level import extract_text
from docx import Document
from openpyxl import load_workbook
from fpdf import FPDF
from config.settings import OUTPUT_DIR, TEMP_DIR


def ensure_directories():
    """Ensure output and temp directories exist."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(TEMP_DIR, exist_ok=True)


def parse_file(file_path):
    """Identifies file type and extracts text accordingly."""
    file_extension = os.path.splitext(file_path)[1].lower()
    try:
        if file_extension == '.pdf':
            return extract_text(file_path)
        elif file_extension == '.docx':
            doc = Document(file_path)
            return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        else:
            raise ValueError("Unsupported file format")
    except Exception as e:
        return {"error": f"Error parsing file: {str(e)}"}


def write_to_excel(data_list, output_file="extracted_resume_data.xlsx"):
    """Writes extracted resume data to an Excel file."""
    ensure_directories()
    output_path = os.path.join(OUTPUT_DIR, output_file)
    
    df = pd.DataFrame(data_list)
    df.to_excel(output_path, index=False, engine='openpyxl')

    # Format Excel columns for better readability
    wb = load_workbook(output_path)
    ws = wb.active

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output_path)
    return output_path


def download_notes_as_pdf(title, content):
    """Creates a downloadable PDF file of the notes."""
    ensure_directories()
    
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    pdf.set_font("Arial", style='B', size=16)
    pdf.cell(200, 10, title, ln=True, align='C')
    pdf.ln(10)  # Line break
    
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, content)

    file_name = f"{title.strip().replace(' ', '_')}.pdf"
    file_path = os.path.join(OUTPUT_DIR, file_name)
    pdf.output(file_path)
    
    return file_path


def create_temp_file(uploaded_file):
    """Creates a temporary file from uploaded file."""
    try:
        with tempfile.NamedTemporaryFile(delete=False, dir=TEMP_DIR) as temp_file:
            temp_file.write(uploaded_file.getbuffer())
            return temp_file.name
    except Exception as e:
        print(f"Error creating temp file: {e}")
        return None


def cleanup_temp_files():
    """Clean up temporary files."""
    try:
        for filename in os.listdir(TEMP_DIR):
            file_path = os.path.join(TEMP_DIR, filename)
            if os.path.isfile(file_path):
                os.unlink(file_path)
    except Exception as e:
        print(f"Error cleaning up temp files: {e}") 